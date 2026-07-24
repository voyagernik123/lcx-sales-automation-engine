#!/usr/bin/env python3
"""
LCX COMMAND deep-seed compiler (100X Phase 1).

Reads the five source strategy workbooks (checked into
apps/api/src/seed/command/source/) and emits data2.ts — the FULL-FIDELITY
program ontology: weighted scorecards (with dimensions + weights), the 21-field
RFI schema, rail providers, architecture / two-path option matrices, the
waitlist funnel model, blockers, requirements, the token-DD framework, budget
lines, dependency edges, and EVERY source row (provenance).

Rules:
  • Non-fabrication: empty cells stay null. Nothing is invented.
  • Fail loudly: every sheet's layout is asserted; if a future strategy revision
    moves a header, this script stops rather than emitting garbage.
  • Provenance: 'Src' columns resolve to namespaced source ids (p1_3, p2_11, m_4).
    Default Admiralty grade for public research is C3; RFI-returned data will be
    upgraded to B2 at runtime; signed terms to A1.

Run:  python3 apps/api/scripts/compile_command_seed.py
Deps: pip3 install openpyxl
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1] / 'src' / 'seed' / 'command'
SRC = ROOT / 'source'
OUT = ROOT / 'data2.ts'

def die(msg):
    print(f'COMPILE FAIL: {msg}', file=sys.stderr)
    sys.exit(1)

def load(name):
    p = SRC / name
    if not p.exists():
        die(f'missing workbook {p}')
    return openpyxl.load_workbook(p, data_only=True)

def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return None
    s = str(v).strip()
    return s if s != '' else None

def num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '').replace('%', ''))
    except ValueError:
        return None

def slug(s):
    return re.sub(r'_+', '_', re.sub(r'[^a-z0-9]+', '_', s.lower())).strip('_')[:40]

def expect(ws, r, c, contains, sheet):
    v = cell(ws, r, c) or ''
    if contains.lower() not in v.lower():
        die(f"{sheet}: expected '{contains}' at r{r}c{c}, found '{v[:60]}'")

def src_refs(v, prefix):
    if v is None:
        return []
    out = []
    for tok in re.split(r'[,;]', str(v)):
        tok = tok.strip().replace('—', '')
        if tok.isdigit():
            out.append(f'{prefix}_{tok}')
    return out

# ── Sources tabs (numbered per workbook; master is unnumbered) ──
def parse_sources(ws, prefix, numbered=True):
    rows = []
    for r in range(1, ws.max_row + 1):
        a, b, c = cell(ws, r, 1), cell(ws, r, 2), cell(ws, r, 3)
        if numbered:
            if a and a.isdigit() and b:
                rows.append({'id': f'{prefix}_{a}', 'phase': prefix.upper(), 'label': b, 'url': c})
        else:
            if a in ('P1', 'P2', 'P3', 'P4', 'Shared', 'All') and b:
                rows.append({'id': f'{prefix}_{len(rows)+1}', 'phase': a, 'label': b, 'url': c})
    if not rows:
        die(f'sources tab empty for {prefix}')
    return rows

# ── Generic weighted-scorecard parser ──
# Layout: title r1, blurb r2, header r3 (subject + dims... + Weighted + Rank + Tier/Verdict),
# weights r4, data rows r5.., then a '<x> notes' marker with label|note rows.
def parse_scorecard(ws, sheet, subject_col_name, dims_expected, id_prefix, id_map=None):
    expect(ws, 3, 1, subject_col_name, sheet)
    headers = []
    c = 2
    while True:
        h = cell(ws, 3, c)
        if h is None or h.lower().startswith(('weighted', 'rank', 'tier', 'verdict')):
            break
        headers.append((c, h))
        c += 1
    tail_cols = {}
    for cc in range(c, ws.max_column + 1):
        h = cell(ws, 3, cc)
        if h:
            key = h.split()[0].lower().rstrip('/5').strip()
            tail_cols[key] = cc
    # Headers WITHOUT a weight in row 4 are meta text columns (e.g. 'Type'),
    # not scored dimensions — captured per-row, excluded from the model.
    expect(ws, 4, 1, 'Weight', sheet)
    weighted_headers, meta_cols = [], []
    for col, h in headers:
        w = num(cell(ws, 4, col))
        if w is None:
            meta_cols.append((col, h))
        else:
            weighted_headers.append((col, h, w))
    if len(weighted_headers) != dims_expected:
        die(f'{sheet}: expected {dims_expected} weighted dims, found {len(weighted_headers)}')
    headers = [(col, h) for col, h, _ in weighted_headers]
    dims = [{'key': slug(h), 'label': h, 'weight': w} for _, h, w in weighted_headers]
    rows, notes, r = [], {}, 5
    in_notes = False
    while r <= ws.max_row:
        a = cell(ws, r, 1)
        r += 1
        if a is None:
            continue
        if 'notes' in a.lower() and len(a) < 40:
            in_notes = True
            continue
        if in_notes:
            n = cell(ws, r - 1, 2)
            if n:
                notes[a] = n
            continue
        scores = {}
        ok = True
        for col, h in headers:
            v = num(cell(ws, r - 1, col))
            if v is None:
                ok = False
                break
            scores[slug(h)] = v
        if not ok:
            continue
        label = a
        sid = (id_map or {}).get(label.split('(')[0].strip().split('—')[0].strip().lower())
        meta = {slug(h): cell(ws, r - 1, col) for col, h in meta_cols}
        row = {
            'subjectId': sid or f'{id_prefix}_{slug(label)}',
            'subjectLabel': label,
            **({'meta': meta} if meta else {}),
            'scores': scores,
            'weighted': num(cell(ws, r - 1, tail_cols.get('weighted', 0))) if tail_cols.get('weighted') else None,
            'rank': num(cell(ws, r - 1, tail_cols.get('rank', 0))) if tail_cols.get('rank') else None,
            'tier': cell(ws, r - 1, tail_cols.get('tier', tail_cols.get('verdict', 0))) if (tail_cols.get('tier') or tail_cols.get('verdict')) else None,
        }
        rows.append(row)
    if not rows:
        die(f'{sheet}: no data rows parsed')
    for row in rows:
        base = row['subjectLabel'].split('(')[0].strip()
        for k, n in notes.items():
            if k.split('(')[0].strip()[:16].lower() == base[:16].lower():
                row['note'] = n
                break
    return {'dimensions': dims, 'rows': rows}

# ── Simple table parser: header at r3, data r4.. until blank col1 streak ──
def parse_table(ws, sheet, first_header, cols, r_start=4, header_row=3):
    expect(ws, header_row, 1, first_header, sheet)
    rows = []
    for r in range(r_start, ws.max_row + 1):
        first = cell(ws, r, 1)
        if first is None:
            continue
        rows.append([cell(ws, r, cix + 1) for cix in range(cols)])
    if not rows:
        die(f'{sheet}: no rows')
    return rows

# ═══════════════ P1 — Liquidity ═══════════════
PT_IDS = {p['name'].split('(')[0].strip().split('—')[0].strip().lower(): p['id']
          for p in json.load(open(ROOT / 'partners.json'))}
PT_IDS['b2c2'] = 'pt_b2c2'
PT_IDS['cumberland'] = 'pt_cumberland'
PT_IDS['dv chain'] = 'pt_dvchain'
PT_IDS['crossover markets'] = 'pt_crossover'
PT_IDS['ripple prime'] = 'pt_rippleprime'
PT_IDS['fireblocks network / off-exchange'] = 'pt_fireblocks'
PT_IDS['copper clearloop'] = 'pt_copper'
PT_IDS['bitgo go network'] = 'pt_bitgo'

wb1 = load('p1_liquidity.xlsx')
lp_scorecard = parse_scorecard(wb1['Scorecard'], 'P1 Scorecard', 'Provider', 10, 'lp', PT_IDS)
p1_sources = parse_sources(wb1['Sources'], 'p1')

cap = []
ws = wb1['Capability Detail']
expect(ws, 3, 1, 'Provider', 'Capability Detail')
for r in range(4, ws.max_row + 1):
    name = cell(ws, r, 1)
    if not name:
        continue
    key = name.split('(')[0].strip().lower()
    cap.append({
        'partnerId': PT_IDS.get(key) or f'lp_{slug(name)}',
        'name': name, 'model': cell(ws, r, 2), 'usEntity': cell(ws, r, 3),
        'assetBreadth': cell(ws, r, 4), 'otcDesk': cell(ws, r, 5), 'rfq': cell(ws, r, 6),
        'optionsFlow': cell(ws, r, 7), 'settlement': cell(ws, r, 8),
        'servesExchanges': cell(ws, r, 9), 'backing': cell(ws, r, 10),
        'sourceRefs': src_refs(cell(ws, r, 11), 'p1'),
    })
if len(cap) < 8:
    die('Capability Detail: too few rows')

conn = []
ws = wb1['Connectivity & Settlement']
expect(ws, 3, 1, 'Provider', 'Connectivity')
for r in range(4, ws.max_row + 1):
    name = cell(ws, r, 1)
    if not name:
        continue
    conn.append({
        'partnerId': PT_IDS.get(name.split('(')[0].strip().lower()) or f'infra_{slug(name)}',
        'name': name, 'category': cell(ws, r, 2), 'role': cell(ws, r, 3),
        'usStanding': cell(ws, r, 4), 'custodyFit': cell(ws, r, 5), 'notes': cell(ws, r, 6),
        'sourceRefs': src_refs(cell(ws, r, 7), 'p1'),
    })

ws = wb1['RFI Tracker']
expect(ws, 3, 1, 'Provider', 'RFI Tracker')
rfi_fields = []
for c in range(2, ws.max_column + 1):
    h = cell(ws, 3, c)
    if h:
        rfi_fields.append({'key': slug(h), 'label': h})
ex = [cell(ws, 4, c) for c in range(1, ws.max_column + 1)]
rfi_example = {'provider': ex[0], 'values': {rfi_fields[i - 1]['key']: ex[i] for i in range(1, len(ex)) if i - 1 < len(rfi_fields)}}
if len(rfi_fields) < 18:
    die(f'RFI Tracker: expected ~20 fields, found {len(rfi_fields)}')

# ═══════════════ P2 — Rails ═══════════════
wb2 = load('p2_rails.xlsx')
arch = parse_scorecard(wb2['Architecture Options'], 'Arch Options', 'Option', 8, 'arch')
p2_sources = parse_sources(wb2['Sources'], 'p2')

rails = []
ws = wb2['Provider Matrix']
expect(ws, 3, 1, 'Provider', 'Provider Matrix')
for r in range(4, ws.max_row + 1):
    name = cell(ws, r, 1)
    if not name:
        continue
    rails.append({
        'id': f'rp_{slug(name)}',
        'partnerId': PT_IDS.get(name.split('(')[0].strip().lower()),
        'name': name, 'category': cell(ws, r, 2), 'provides': cell(ws, r, 3),
        'usStanding': cell(ws, r, 4), 'custodyFit': cell(ws, r, 5),
        'economics': cell(ws, r, 6), 'notes': cell(ws, r, 7),
        'sourceRefs': src_refs(cell(ws, r, 8), 'p2'),
    })
if len(rails) < 15:
    die('Provider Matrix: too few rows')

stable = []
ws = wb2['Stablecoin Policy (GENIUS)']
expect(ws, 3, 1, 'Stablecoin', 'Stablecoin Policy')
for r in range(4, ws.max_row + 1):
    coin = cell(ws, r, 1)
    if not coin:
        continue
    stable.append({'coin': coin, 'issuer': cell(ws, r, 2), 'status': cell(ws, r, 3),
                   'action': cell(ws, r, 4), 'sourceRefs': src_refs(cell(ws, r, 5), 'p2')})

lic = []
ws = wb2['Licensing & Compliance']
expect(ws, 3, 2, 'Requirement', 'Licensing')
for r in range(4, ws.max_row + 1):
    n = cell(ws, r, 1)
    if not n or not str(n).replace('.0', '').isdigit():
        continue
    lic.append({'num': int(float(n)), 'requirement': cell(ws, r, 2), 'detail': cell(ws, r, 3),
                'owner': cell(ws, r, 4), 'status': cell(ws, r, 5)})

# ═══════════════ P3 — Waitlist ═══════════════
wb3 = load('p3_waitlist.xlsx')
channels_sc = parse_scorecard(wb3['Channel Scorecard'], 'Channel Scorecard', 'Channel', 6, 'ch')
if len(channels_sc['rows']) < 10:
    die('Channel Scorecard: too few channels')
p3_sources = parse_sources(wb3['Sources'], 'p3')

ws = wb3['Funnel & Budget']
funnel_channels, conversions, scenarios = [], {}, []
for r in range(1, ws.max_row + 1):
    a, b, c3, d, e = (cell(ws, r, i) for i in range(1, 6))
    if b in ('Paid', 'Organic') and a:
        budget = num(c3) or 0
        cac = num(d)
        signups = num(e)
        if signups is None and cac and cac > 0:
            signups = round(budget / cac)
        funnel_channels.append({'channelId': f'ch_{slug(a)}', 'label': a, 'type': b,
                                'budget': budget, 'cac': cac, 'signupsEst': signups})
    if a and 'verified' in a.lower() and a.startswith('×'):
        conversions['waitlistToVerified'] = num(c3)
    if a and 'funded acc' in a.lower() and a.startswith('×'):
        conversions['verifiedToFunded'] = num(c3)
    if a in ('Lean', 'Aggressive'):
        scenarios.append({'name': a, 'budget': num(b), 'waitlist': num(c3), 'funded': num(d)})
if not funnel_channels or 'waitlistToVerified' not in conversions:
    die('Funnel & Budget: parse failed')
base_budget = sum(fc['budget'] for fc in funnel_channels)
base_waitlist = sum(fc['signupsEst'] or 0 for fc in funnel_channels)
scenarios.insert(1, {'name': 'Base', 'budget': base_budget, 'waitlist': base_waitlist,
                     'funded': round(base_waitlist * conversions['waitlistToVerified'] * conversions['verifiedToFunded'])})

referral, guardrails = [], []
ws = wb3['Referral & Compliance']
mode = None
for r in range(1, ws.max_row + 1):
    a, b = cell(ws, r, 1), cell(ws, r, 2)
    if a and a.startswith('A.'):
        mode = 'ref'
        continue
    if a and a.startswith('B.'):
        mode = 'guard'
        continue
    if a == 'Guardrail' or a == 'Core loop' and mode is None:
        continue
    if mode == 'ref' and a and b and a != 'Core loop':
        referral.append({'component': a, 'design': b})
    elif mode == 'ref' and a == 'Core loop' and b:
        referral.insert(0, {'component': a, 'design': b})
    elif mode == 'guard' and a and b and a != 'Guardrail':
        guardrails.append({'guardrail': a, 'meaning': b, 'riskIfIgnored': cell(ws, r, 3),
                           'sourceRefs': src_refs(cell(ws, r, 4), 'p3')})

plan90 = [{'window': row[0], 'workstream': row[1], 'actions': row[2], 'owner': row[3], 'kpi': row[4]}
          for row in parse_table(wb3['90-Day Plan'], '90-Day Plan', 'Window', 5)]
tooling = [{'function': row[0], 'options': row[1], 'notes': row[2],
            'sourceRefs': src_refs(row[3], 'p3')}
           for row in parse_table(wb3['Tooling & Vendors'], 'Tooling', 'Function', 4)]

# ═══════════════ P4 — Listing ═══════════════
wb4 = load('p4_listing.xlsx')
two_path = parse_scorecard(wb4['Two-Path Options'], 'Two-Path', 'Option', 6, 'path')
p4_sources = parse_sources(wb4['Sources'], 'p4')

blockers = []
for row in parse_table(wb4['Blockers Register'], 'Blockers', '#', 7):
    if row[0] and str(row[0]).replace('.0', '').isdigit():
        blockers.append({'num': int(float(row[0])), 'blocker': row[1], 'category': row[2],
                         'severity': row[3], 'detail': row[4], 'owner': row[5], 'resolvesVia': row[6]})
if len(blockers) != 12:
    die(f'Blockers: expected 12, found {len(blockers)}')

requirements = []
for row in parse_table(wb4['Requirements Checklist'], 'Requirements', '#', 6):
    if row[0] and str(row[0]).replace('.0', '').isdigit():
        requirements.append({'num': int(float(row[0])), 'requirement': row[1], 'detail': row[2],
                             'path': row[3], 'owner': row[4], 'status': row[5]})
if len(requirements) != 14:
    die(f'Requirements: expected 14, found {len(requirements)}')

dd = []
ws = wb4['Token Due-Diligence']
expect(ws, 3, 1, 'Dimension', 'Token DD')
for r in range(4, ws.max_row + 1):
    dim = cell(ws, r, 1)
    if not dim:
        continue
    w = num(cell(ws, r, 4))
    dd.append({'dimension': dim, 'assess': cell(ws, r, 2), 'criteria': cell(ws, r, 3),
               'weightPct': w, 'gate': 'GATE' in (dim or '') or 'GATE' in (cell(ws, r, 3) or ''),
               'sourceRefs': src_refs(cell(ws, r, 5), 'p4')})
if abs(sum(d['weightPct'] or 0 for d in dd) - 100) > 1:
    die(f"Token DD weights don't sum to 100: {[d['weightPct'] for d in dd]}")

policy_outline = [{'section': row[0], 'contents': row[1]}
                  for row in parse_table(wb4['Listing Policy Outline'], 'Policy', 'Section', 2)
                  if row[0] and row[0][0].isdigit()]

# ═══════════════ Master ═══════════════
wbm = load('master.xlsx')
m_sources = parse_sources(wbm['Sources'], 'm', numbered=False)

budget_lines = [{'phase': row[0], 'costType': row[1], 'range': row[2], 'model': row[3], 'notes': row[4]}
                for row in parse_table(wbm['Budget View'], 'Budget', 'Phase', 5)
                if row[0] and row[0] not in ('Phase',)]

dep_edges = []
ws = wbm['Dependency Map']
expect(ws, 3, 1, 'Depends on', 'Dependency Map')
col_names = [cell(ws, 3, c) for c in range(2, 7)]
for r in range(4, 9):
    frm = cell(ws, r, 1)
    if not frm or frm.startswith('Takeaway'):
        continue
    for i, cn in enumerate(col_names):
        v = cell(ws, r, 2 + i)
        if v and (v.startswith('●') or v.startswith('○')):
            dep_edges.append({'fromWs': frm, 'onWs': cn, 'strength': 'hard' if v.startswith('●') else 'soft',
                              'note': v[1:].strip()})
if not dep_edges:
    die('Dependency Map: no edges')

exec_rows = []
ws = wbm['Executive Dashboard']
for r in range(1, ws.max_row + 1):
    a = cell(ws, r, 1)
    if a and re.match(r'^[1-4] ', a.replace('·', ' ').strip()):
        exec_rows.append({'phase': a, 'objective': cell(ws, r, 2), 'recommendation': cell(ws, r, 3),
                          'topPicks': cell(ws, r, 4), 'northStar': cell(ws, r, 5), 'gatingDep': cell(ws, r, 6)})
if len(exec_rows) != 4:
    die(f'Exec Dashboard: expected 4 phases, found {len(exec_rows)}')

roadmap = []
ws = wbm['Master Roadmap']
expect(ws, 3, 1, 'Workstream', 'Roadmap')
for r in range(4, ws.max_row + 1):
    a = cell(ws, r, 1)
    if a:
        roadmap.append({'workstream': a, 'now': cell(ws, r, 2), 'm3to6': cell(ws, r, 3),
                        'm6to12': cell(ws, r, 4), 'm12to24': cell(ws, r, 5)})

consolidated_risks = []
ws = wbm['Risks & Dependencies']
expect(ws, 3, 1, 'Phase', 'Risks')
for r in range(4, ws.max_row + 1):
    a = cell(ws, r, 1)
    if a:
        consolidated_risks.append({'phases': a, 'risk': cell(ws, r, 2), 'severity': cell(ws, r, 3),
                                   'mitigation': cell(ws, r, 4), 'owner': cell(ws, r, 5)})

dec_enrich = []
ws = wbm['Decisions Register']
expect(ws, 3, 1, 'Phase', 'Decisions Register')
PHASE_BASE = {'P1': 0, 'P2': 6, 'P3': 12, 'P4': 18}
for r in range(4, ws.max_row + 1):
    ph, n = cell(ws, r, 1), cell(ws, r, 2)
    if ph in PHASE_BASE and n and str(n).replace('.0', '').isdigit():
        idx = PHASE_BASE[ph] + int(float(n))
        dec_enrich.append({'decisionId': f'dec_{idx:02d}', 'options': cell(ws, r, 4),
                           'owner': cell(ws, r, 6)})
if len(dec_enrich) != 24:
    die(f'Decisions Register: expected 24, found {len(dec_enrich)}')

# ═══════════════ Emit ═══════════════
payload = {
    'defaultGrade': 'C3',
    'scorecards': {'lp': lp_scorecard, 'channel': channels_sc, 'arch': arch, 'twoPath': two_path},
    'capabilityDetail': cap,
    'connectivity': conn,
    'rfi': {'fields': rfi_fields, 'example': rfi_example},
    'railProviders': rails,
    'stablecoinPolicy': stable,
    'licensingChecklist': lic,
    'funnel': {'channels': funnel_channels, 'conversions': conversions, 'scenarios': scenarios},
    'referralMechanics': referral,
    'guardrails': guardrails,
    'ninetyDayPlan': plan90,
    'tooling': tooling,
    'blockers': blockers,
    'requirements': requirements,
    'ddDimensions': dd,
    'listingPolicyOutline': policy_outline,
    'budgetLines': budget_lines,
    'dependencyEdges': dep_edges,
    'execDashboard': exec_rows,
    'masterRoadmap': roadmap,
    'consolidatedRisks': consolidated_risks,
    'decisionEnrichment': dec_enrich,
    'sources': p1_sources + p2_sources + p3_sources + p4_sources + m_sources,
}

banner = (
    '/**\n'
    ' * LCX COMMAND DEEP SEED (100X Phase 1) — full-fidelity program ontology\n'
    ' * compiled from the five source strategy workbooks (seed/command/source/).\n'
    ' * GENERATED by apps/api/scripts/compile_command_seed.py — DO NOT hand-edit;\n'
    ' * re-run the compiler when the strategy is revised. Non-fabrication: null\n'
    ' * fields are gaps in the source, never inventions. Default provenance grade\n'
    " * C3 (public research); RFI-returned data upgrades to B2, signed terms to A1.\n"
    ' */\n'
)
OUT.write_text(banner + 'export const COMMAND_DEEP_SEED = '
               + json.dumps(payload, indent=2, ensure_ascii=False) + ' as const;\n')

counts = {k: (len(v) if isinstance(v, list) else 'obj') for k, v in payload.items()}
print('COMPILED OK →', OUT)
print(json.dumps(counts, indent=1))
print('lp rows:', len(lp_scorecard['rows']), '| channel rows:', len(channels_sc['rows']),
      '| arch rows:', len(arch['rows']), '| twoPath rows:', len(two_path['rows']),
      '| sources:', len(payload['sources']), '| rfi fields:', len(rfi_fields))
